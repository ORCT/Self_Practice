from bs4 import BeautifulSoup as BS
import urllib.request as req
import openpyxl
import os
from openpyxl.drawing.image import Image

if not os.path.exists('./melon_chart.xlsx'):
    book = openpyxl.Workbook()
    book.save('./melon_chart.xlsx')

if not os.path.exists('./album_image'):
    os.mkdir('./album_image')

headders = req.Request('https://www.melon.com/chart/', headers={'User-Agent':'Mozilla/5.0'})
code = req.urlopen(headders)
soup = BS(code, 'html.parser')
title = soup.select('div.ellipsis.rank01 a')
name = soup.select('div.ellipsis.rank02 span.checkEllipsis')
album = soup.select('div.ellipsis.rank03 a')
image = soup.select('a.image_typeAll img')

book = openpyxl.load_workbook('./melon_chart.xlsx')
#book['Sheet1']
sheet = book.active # 자동으로 열리는 시트
sheet.column_dimensions['A'].width = '16.2'
sheet.column_dimensions['B'].width = '50'
sheet.column_dimensions['C'].width = '30'
sheet.column_dimensions['D'].width = '35'
row_num = 1

for i in range(len(title)):
    req.urlretrieve(image[i].attrs['src'], './album_image/{}.png'.format(row_num))
    print(title[i].string,name[i].text,album[i].string,image[i].attrs['src'])
    image_for_excel = Image('./album_image/{}.png'.format(row_num))
    sheet.row_dimensions[row_num].height = 95
    sheet.add_image(image_for_excel, 'A{}'.format(row_num))
    sheet.cell(row=row_num,column=2).value = title[i].string
    sheet.cell(row=row_num,column=3).value = name[i].text
    sheet.cell(row=row_num,column=4).value = album[i].string
    book.save('./melon_chart.xlsx')
    row_num += 1